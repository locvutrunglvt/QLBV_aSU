import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib

def create_qlbv_template():
    wb = openpyxl.Workbook()
    # Ensure Author metadata is set strictly to "Lộc Vũ Trung"
    wb.properties.creator = "Lộc Vũ Trung"
    wb.properties.lastModifiedBy = "Lộc Vũ Trung"
    wb.properties.title = "Hệ thống Quản lý Dữ liệu Bảo trì Hàng rào điện (QLBV)"
    wb.properties.subject = "Dự án Bảo tồn Voi TP. Đồng Nai"
    wb.properties.description = "Mẫu CSDL Google Sheets cho WebApp QLBV"

    # Default sheet
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    # Styling helper functions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="595959")
    bold_font = Font(name="Calibri", size=11, bold=True)
    card_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # 1. SHEET DASHBOARD (TỔNG HỢP & THỐNG KÊ QUA CÔNG THỨC)
    # -------------------------------------------------------------
    ws_dash['A1'] = "HỆ THỐNG QUẢN LÝ BẢO TRÌ & VẬN HÀNH HÀNG RÀO ĐIỆN (QLBV)"
    ws_dash['A1'].font = title_font
    ws_dash['A2'] = "Dự án Khẩn cấp Bảo tồn Voi TP. Đồng Nai (QĐ 704/QĐ-SNNMT & 595/QĐ-SNNMT)"
    ws_dash['A2'].font = subtitle_font

    ws_dash['A4'] = "BẢNG TỔNG HỢP SỐ LIỆU THỰC HIỆN HIỆN TRƯỜNG"
    ws_dash['A4'].font = bold_font

    dash_headers = ["Hạng mục thống kê", "Số lượng / Giá trị", "Đơn vị tính", "Ghi chú"]
    for c_idx, h in enumerate(dash_headers, 1):
        cell = ws_dash.cell(row=5, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    dash_rows = [
        ("Tổng số lượt lập Biên bản Bảo trì sự cố", '=COUNTA(BB_BaoTri!A2:A1000)', "Lượt", "Liên kết từ sheet BB_BaoTri"),
        ("Tổng số lượt lập Biên bản Vệ sinh phát dọn", '=COUNTA(BB_VeSinh!A2:A1000)', "Lượt", "Liên kết từ sheet BB_VeSinh"),
        ("Tổng chiều dài tuyến đã phát dọn vệ sinh", '=SUM(BB_VeSinh!L2:L1000)', "Mét", "Tổng cộng công thức =SUM"),
        ("Tổng diện tích thực bì đã phát dọn", '=SUM(BB_VeSinh!M2:M1000)', "m²", "Tổng cộng công thức =SUM"),
        ("Tổng diện tích quy đổi ra Hecta (ha)", '=B9/10000', "ha", "Công thức quy đổi =B9/10000"),
        ("Tổng số ngày ghi Nhật ký thi công", '=COUNTA(NhatKy_ThiCong!A2:A1000)', "Ngày", "Liên kết từ sheet NhatKy_ThiCong"),
        ("Tổng số tài khoản người dùng trong hệ thống", '=COUNTA(Users!A2:A100)', "Tài khoản", "Liên kết từ sheet Users")
    ]

    for r_idx, (item, formula, unit, note) in enumerate(dash_rows, 6):
        c1 = ws_dash.cell(row=r_idx, column=1, value=item)
        c2 = ws_dash.cell(row=r_idx, column=2, value=formula)
        c3 = ws_dash.cell(row=r_idx, column=3, value=unit)
        c4 = ws_dash.cell(row=r_idx, column=4, value=note)
        
        c2.font = bold_font
        c2.alignment = Alignment(horizontal="right")
        c3.alignment = Alignment(horizontal="center")
        for c in [c1, c2, c3, c4]:
            c.border = thin_border

    # -------------------------------------------------------------
    # 2. SHEET USERS (TÀI KHOẢN & PHÂN QUYỀN)
    # -------------------------------------------------------------
    ws_users = wb.create_sheet(title="Users")
    user_headers = ["user_id", "username", "password_hash", "full_name", "role", "organization", "phone", "status", "last_login"]
    for c_idx, h in enumerate(user_headers, 1):
        cell = ws_users.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Mật khẩu mặc định: 123456 (SHA-256: 8d969eef6ecad3c29a3a629280e686cf0c3f5d5a86aff3ca12020c923adc6c92)
    sample_users = [
        ("U001", "admin", "8d969eef6ecad3c29a3a629280e686cf0c3f5d5a86aff3ca12020c923adc6c92", "Lộc Vũ Trung (Admin)", "Admin", "Ban Quản lý / Chi cục Kiểm lâm", "0900000001", "Active", ""),
        ("U002", "giamsat", "8d969eef6ecad3c29a3a629280e686cf0c3f5d5a86aff3ca12020c923adc6c92", "Cán bộ Giám sát Kiểm lâm", "Supervisor", "Hạt Kiểm lâm Khu vực", "0900000002", "Active", ""),
        ("U003", "baotri01", "8d969eef6ecad3c29a3a629280e686cf0c3f5d5a86aff3ca12020c923adc6c92", "Tổ trưởng Đội Bảo trì 1", "Worker", "Đơn vị Bảo trì HRD", "0900000003", "Active", ""),
        ("U004", "baotri02", "8d969eef6ecad3c29a3a629280e686cf0c3f5d5a86aff3ca12020c923adc6c92", "Tổ trưởng Đội Bảo trì 2", "Worker", "Đơn vị Bảo trì HRD", "0900000004", "Active", "")
    ]
    for r_idx, u in enumerate(sample_users, 2):
        for c_idx, val in enumerate(u, 1):
            cell = ws_users.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

    # -------------------------------------------------------------
    # 3. SHEET MASTER DATA (VALIDATION / GỢI Ý CHỌN NHANH)
    # -------------------------------------------------------------
    ws_master = wb.create_sheet(title="MasterData")
    master_headers = ["category", "code", "name", "extra_info"]
    for c_idx, h in enumerate(master_headers, 1):
        cell = ws_master.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    master_items = [
        # Tuyến
        ("Tuyen", "T01", "Tuyến Hàng rào điện Định Quán (Đoạn 1)", "Trụ 01 - Trụ 150"),
        ("Tuyen", "T02", "Tuyến Hàng rào điện Vĩnh Cửu (Đoạn 2)", "Trụ 151 - Trụ 320"),
        ("Tuyen", "T03", "Tuyến Hàng rào điện Tân Phú (Đoạn 3)", "Trụ 321 - Trụ 480"),
        ("Tuyen", "T04", "Tuyến Hàng rào điện Bù Gia Mập giáp ranh", "Trụ 481 - Trụ 600"),
        # Đơn vị bảo trì
        ("DonViBaoTri", "DV01", "Công ty TNHH Kỹ thuật Xây dựng & Môi trường Đồng Nai", "Đơn vị trúng thầu bảo trì"),
        ("DonViBaoTri", "DV02", "Tổ Kỹ thuật Bảo trì Hàng rào điện Chi cục Kiểm lâm", "Đội cơ động"),
        # Chủ rừng
        ("ChuRung", "CR01", "Ban Quản lý Rừng phòng hộ Tân Phú", "Huyện Tân Phú, Đồng Nai"),
        ("ChuRung", "CR02", "Khu Bảo tồn Thiên nhiên - Văn hóa Đồng Nai", "Huyện Vĩnh Cửu, Đồng Nai"),
        ("ChuRung", "CR03", "Công ty TNHH MTV Lâm nghiệp La Ngà", "Huyện Định Quán, Đồng Nai"),
        # Kiểm lâm
        ("KiemLam", "KL01", "Hạt Kiểm lâm Huyện Định Quán", "Chi cục Kiểm lâm Đồng Nai"),
        ("KiemLam", "KL02", "Hạt Kiểm lâm Huyện Vĩnh Cửu", "Chi cục Kiểm lâm Đồng Nai"),
        ("KiemLam", "KL03", "Hạt Kiểm lâm Khu Bảo tồn Đồng Nai", "Chi cục Kiểm lâm Đồng Nai"),
        # Sự cố thường gặp
        ("SuCo", "SC01", "Đứt dây dẫn xung điện do cây đổ/gãy đè lên", "Cần nối dây & phát dọn cây"),
        ("SuCo", "SC02", "Sét đánh gây hỏng bộ tạo xung (Energizer)", "Cần thay bo mạch/máy phát"),
        ("SuCo", "SC03", "Bình ắc quy / Tấm pin năng lượng mặt trời hỏng, yếu điện", "Cần kiểm tra sạc/ắc quy"),
        ("SuCo", "SC04", "Sứ cách điện bị vỡ/nứt làm rò điện xuống đất", "Cần thay sứ cách điện"),
        ("SuCo", "SC05", "Trụ rào bị nghiêng, gãy do voi/gia súc va quẹt", "Cần dựng lại/trồng trụ mới"),
        ("SuCo", "SC06", "Cỏ cây mọc um tùm chạm vào dây gây ngắn mạch", "Cần phát dọn thực bì"),
        # Thiết bị vật tư
        ("ThietBi", "TB01", "Bộ phát xung điện (Energizer) 12V/220V", "Cái"),
        ("ThietBi", "TB02", "Bình ắc quy lưu điện 12V - 100Ah", "Bình"),
        ("ThietBi", "TB03", "Tấm pin năng lượng mặt trời Solar 100W/150W", "Tấm"),
        ("ThietBi", "TB04", "Sứ cách điện néo / Sứ đỡ dây (Insulator)", "Cái"),
        ("ThietBi", "TB05", "Dây cáp thép bọc kẽm chịu lực 2.5mm / 3.0mm", "Mét"),
        ("ThietBi", "TB06", "Bộ chống sét lan truyền cho hàng rào điện", "Bộ"),
        ("ThietBi", "TB07", "Bộ đo điện áp hàng rào chuyên dụng (Fault Finder)", "Cái"),
        ("ThietBi", "TB08", "Trụ bê tông / Trụ sắt gia cố hàng rào", "Trụ"),
        # Chủ đầu tư & Giám sát
        ("ChuDauTu", "CDT01", "Chi cục Kiểm lâm Tỉnh Đồng Nai", "0251.3822xxx"),
        ("TuVanGiamSat", "TVGS01", "Ban Giám sát & Quản lý Dự án Bảo tồn Voi", "0912.345xxx")
    ]

    for r_idx, item in enumerate(master_items, 2):
        for c_idx, val in enumerate(item, 1):
            cell = ws_master.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

    # -------------------------------------------------------------
    # 4. SHEET BB_BAOTRI (BIÊN BẢN BẢO TRÌ)
    # -------------------------------------------------------------
    ws_baotri = wb.create_sheet(title="BB_BaoTri")
    bt_headers = [
        "record_id", "created_at", "created_by", "ngay_lap", "gio_lap", "dia_diem", "tuyen_hrd",
        "dai_dien_baotri", "dai_dien_churung", "dai_dien_kiemlam", "su_co", "nguyen_nhan",
        "thiet_bi_hong", "thiet_bi_thay_the", "ket_qua_khac_phuc", "tinh_trang", "photos_url", "signature", "notes"
    ]
    for c_idx, h in enumerate(bt_headers, 1):
        cell = ws_baotri.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # -------------------------------------------------------------
    # 5. SHEET BB_VESINH (BIÊN BẢN VỆ SINH PHÁT DỌN)
    # -------------------------------------------------------------
    ws_vesinh = wb.create_sheet(title="BB_VeSinh")
    vs_headers = [
        "record_id", "created_at", "created_by", "ngay_lap", "gio_lap", "dia_diem", "tuyen_hrd",
        "dai_dien_baotri", "dai_dien_churung", "ds_nhan_cong", "chi_tiet_doan_ve_sinh",
        "tong_chieu_dai_m", "tong_dien_tich_m2", "danh_gia_ket_qua", "photos_url", "signature", "notes"
    ]
    for c_idx, h in enumerate(vs_headers, 1):
        cell = ws_vesinh.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # -------------------------------------------------------------
    # 6. SHEET NHATKY_THICONG (NHẬT KÝ CÔNG TRƯỜNG)
    # -------------------------------------------------------------
    ws_nhatky = wb.create_sheet(title="NhatKy_ThiCong")
    nk_headers = [
        "record_id", "created_at", "created_by", "ngay_ghi", "thoi_tiet", "hang_muc", "dia_diem",
        "chu_dau_tu", "tu_van_giam_sat", "nha_thau_thi_cong", "ds_thiet_bi", "ds_nhan_cong",
        "noi_dung_cong_viec", "danh_gia_giam_sat", "photos_url", "signature", "notes"
    ]
    for c_idx, h in enumerate(nk_headers, 1):
        cell = ws_nhatky.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # -------------------------------------------------------------
    # 7. SHEET SYSTEM LOGS
    # -------------------------------------------------------------
    ws_logs = wb.create_sheet(title="SystemLogs")
    log_headers = ["timestamp", "username", "action", "details"]
    for c_idx, h in enumerate(log_headers, 1):
        cell = ws_logs.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        sheet.views.sheetView[0].showGridLines = True
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='): # Formula length estimate
                    max_len = max(max_len, 15)
                else:
                    max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output_path = r"g:\My Drive\GIS\Out\A Su K40 QLBV\backend\QLBV_GoogleSheets_Template.xlsx"
    wb.save(output_path)
    print("SUCCESS: Template created at " + output_path)

if __name__ == "__main__":
    create_qlbv_template()
